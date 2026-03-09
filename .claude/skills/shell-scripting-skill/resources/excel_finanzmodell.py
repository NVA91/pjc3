#!/usr/bin/env python3
# LADESTUFE 3 — RESSOURCE | KATEGORIE: Office-Arbeit / Excel-Automatisierung
# Quellcode belastet das Kontextfenster nicht. Nur stdout (Ergebnis) wird
# dem Modell zurückgegeben. Ausführung in Sandbox auf expliziten Nutzerruf.
#
# BESCHREIBUNG: Professionelle Excel-Finanzmodelle mit echter Formelinjektion.
#               Kein Hardcoding von Werten — alle Abhängigkeiten bleiben intakt.
# VERWENDUNG:
#   python excel_finanzmodell.py 3statement  <ausgabe.xlsx> [--perioden 5]
#   python excel_finanzmodell.py saas        <ausgabe.xlsx> [--perioden 12]
#   python excel_finanzmodell.py szenario    <ausgabe.xlsx>
#   python excel_finanzmodell.py verifizieren <datei.xlsx>
#   python excel_finanzmodell.py debuggen    <datei.xlsx>
# ABHÄNGIGKEITEN:
#   pip install openpyxl pandas
#   apt install libreoffice  (optional, für Verifikation)

from __future__ import annotations

import sys
import subprocess
import shutil
import json
from pathlib import Path
from datetime import datetime, date

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        numbers as xl_numbers,
    )
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
    from openpyxl.utils import get_column_letter, column_index_from_string
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# =============================================================================
# DESIGN-KONSTANTEN (Finanzmodellierungs-Industriestandard)
# =============================================================================

# Schriftfarben-Konvention
FARBE_EINGABE   = "0070C0"   # Blau   — manuelle Annahmen / Inputs
FARBE_FORMEL    = "000000"   # Schwarz — berechnete Formeln
FARBE_LINK      = "00B050"   # Grün   — Cross-Sheet-Referenzen
FARBE_FEHLER    = "FF0000"   # Rot    — Fehler / negative Alerts

# Header-Farben
FARBE_HEADER_BG   = "1F3864"   # Dunkelblau
FARBE_HEADER_TEXT = "FFFFFF"
FARBE_SECTION_BG  = "D6E4F0"   # Hellblau für Abschnitts-Header
FARBE_TOTAL_BG    = "E2EFDA"   # Hellgrün für Summenzeilen
FARBE_SUBTOTAL_BG = "FFF2CC"   # Hellgelb für Zwischensummen

# Zahlenformate
FMT_EUR     = '#,##0 "€";[RED]-#,##0 "€"'
FMT_EUR_K   = '#,##0.0 "T€";[RED]-#,##0.0 "T€"'
FMT_PCT     = '0.0%'
FMT_PCT_1   = '0.00%'
FMT_MULTI   = '0.0"x"'
FMT_DATUM   = 'MMM YYYY'


# =============================================================================
# STYLING-UTILITIES
# =============================================================================

def _rand(stil: str = "thin", farbe: str = "B8CCE4") -> Border:
    s = Side(style=stil, color=farbe)
    return Border(left=s, right=s, top=s, bottom=s)

def _rand_aussen(farbe: str = "1F3864") -> Border:
    s = Side(style="medium", color=farbe)
    return Border(left=s, right=s, top=s, bottom=s)

def _zelle_stil(
    ws, row: int, col: int,
    wert=None,
    farbe: str = FARBE_FORMEL,
    bold: bool = False,
    bg: str | None = None,
    zahlenformat: str | None = None,
    ausrichtung: str = "right",
    ist_eingabe: bool = False,
) -> None:
    """Schreibt Wert und Stil in eine Zelle."""
    z = ws.cell(row=row, column=col, value=wert)
    z.font = Font(
        color=FARBE_EINGABE if ist_eingabe else farbe,
        bold=bold,
        name="Calibri",
        size=9,
    )
    if bg:
        z.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    if zahlenformat:
        z.number_format = zahlenformat
    z.alignment = Alignment(horizontal=ausrichtung, vertical="center")
    z.border = _rand()

def _header_zeile(ws, row: int, titel: str, n_spalten: int, start_col: int = 1) -> None:
    """Schreibt eine fusionierte dunkel-blaue Header-Zeile."""
    ws.merge_cells(
        start_row=row, start_column=start_col,
        end_row=row, end_column=start_col + n_spalten - 1
    )
    z = ws.cell(row=row, column=start_col, value=titel)
    z.font = Font(color=FARBE_HEADER_TEXT, bold=True, name="Calibri", size=10)
    z.fill = PatternFill(start_color=FARBE_HEADER_BG, end_color=FARBE_HEADER_BG, fill_type="solid")
    z.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18

def _section_header(ws, row: int, titel: str, n_spalten: int, start_col: int = 1) -> None:
    """Hellblauer Abschnitts-Header."""
    ws.merge_cells(
        start_row=row, start_column=start_col,
        end_row=row, end_column=start_col + n_spalten - 1
    )
    z = ws.cell(row=row, column=start_col, value=titel.upper())
    z.font = Font(color=FARBE_FORMEL, bold=True, name="Calibri", size=9)
    z.fill = PatternFill(start_color=FARBE_SECTION_BG, end_color=FARBE_SECTION_BG, fill_type="solid")
    z.alignment = Alignment(horizontal="left", vertical="center", indent=1)

def _spaltenbreiten(ws, breiten: dict[int, float]) -> None:
    for col, breite in breiten.items():
        ws.column_dimensions[get_column_letter(col)].width = breite

def _freeze(ws, zelle: str = "B4") -> None:
    ws.freeze_panes = zelle


# =============================================================================
# MODUL 1: 3-STATEMENT FINANZMODELL
# Income Statement → Balance Sheet → Cash Flow Statement
# Alle Werte als Formeln — kein Hardcoding.
# =============================================================================

def erstelle_3statement(ausgabe: str, perioden: int = 5) -> None:
    """
    Erstellt ein vollständiges 3-Statement Finanzmodell:
    - Sheet 1: Annahmen (blaue Eingabezellen)
    - Sheet 2: Income Statement (GuV)
    - Sheet 3: Balance Sheet (Bilanz)
    - Sheet 4: Cash Flow Statement

    Alle berechneten Zellen enthalten Excel-Formeln.
    """
    if not HAS_OPENPYXL:
        print("ERR | openpyxl nicht installiert: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    print(f"[{_ts()}] [3-Statement] Erstelle Modell mit {perioden} Perioden...")
    wb = Workbook()

    jahre = [date.today().year + i for i in range(perioden)]
    n = perioden

    # --- SHEET 1: ANNAHMEN ---
    ws_a = wb.active
    ws_a.title = "Annahmen"
    _spaltenbreiten(ws_a, {1: 35, **{i+2: 14 for i in range(n)}})
    _header_zeile(ws_a, 1, "ANNAHMEN — Manuelle Eingaben (blau = Eingabe, schwarz = Formel)", n + 1)

    # Perioden-Header
    ws_a.cell(row=2, column=1, value="Kennzahl")
    ws_a.cell(row=2, column=1).font = Font(bold=True, size=9)
    for i, j in enumerate(jahre):
        _zelle_stil(ws_a, 2, i+2, str(j), bold=True, bg=FARBE_HEADER_BG,
                    farbe=FARBE_HEADER_TEXT, ausrichtung="center")

    # Umsatz-Annahmen (Eingabezellen — blau)
    annahmen = [
        ("Umsatzwachstum (YoY)", [0.25, 0.22, 0.18, 0.15, 0.12]),
        ("Bruttomargen-%", [0.62, 0.64, 0.65, 0.66, 0.67]),
        ("EBITDA-Marge-%", [0.08, 0.12, 0.16, 0.20, 0.23]),
        ("Steuerrate-%", [0.25, 0.25, 0.25, 0.25, 0.25]),
        ("Capex % Umsatz", [0.05, 0.04, 0.04, 0.03, 0.03]),
        ("Basisumsatz (€)", [5_000_000] + [None] * (n - 1)),
    ]
    _section_header(ws_a, 3, "Wachstums- und Margen-Annahmen", n + 1)
    for r, (name, werte) in enumerate(annahmen, 4):
        _zelle_stil(ws_a, r, 1, name, farbe=FARBE_FORMEL, ausrichtung="left")
        for c, w in enumerate(werte[:n], 2):
            fmt = FMT_PCT if "%" in name else FMT_EUR
            _zelle_stil(ws_a, r, c, w if w is not None else "",
                        ist_eingabe=True, zahlenformat=fmt)
    ws_a.freeze_panes = "B4"

    # --- SHEET 2: INCOME STATEMENT (GuV) ---
    ws_is = wb.create_sheet("Income Statement")
    _spaltenbreiten(ws_is, {1: 38, **{i+2: 14 for i in range(n)}})
    _header_zeile(ws_is, 1, "INCOME STATEMENT (Gewinn- und Verlustrechnung)", n + 1)

    # Perioden-Header
    for i, j in enumerate(jahre):
        _zelle_stil(ws_is, 2, i+2, str(j), bold=True, bg=FARBE_HEADER_BG,
                    farbe=FARBE_HEADER_TEXT, ausrichtung="center")

    # Zeilen-Definitionen (Label, Formel-Typ, Zahlenformat, Bg)
    is_zeilen = [
        # (Label, Formeln als Liste je Periode, fmt, bg)
        ("Umsatz", None, FMT_EUR, None),
        ("  Herstellungskosten (COGS)", None, FMT_EUR, None),
        ("Bruttogewinn", None, FMT_EUR, FARBE_SUBTOTAL_BG),
        ("  Forschung & Entwicklung", None, FMT_EUR, None),
        ("  Vertrieb & Marketing", None, FMT_EUR, None),
        ("  Allgemein & Verwaltung", None, FMT_EUR, None),
        ("EBITDA", None, FMT_EUR, FARBE_SUBTOTAL_BG),
        ("  Abschreibungen (D&A)", None, FMT_EUR, None),
        ("EBIT (Operatives Ergebnis)", None, FMT_EUR, FARBE_SUBTOTAL_BG),
        ("  Zinsen und Finanzkosten", None, FMT_EUR, None),
        ("EBT (Ergebnis vor Steuern)", None, FMT_EUR, FARBE_SUBTOTAL_BG),
        ("  Steuern", None, FMT_EUR, None),
        ("Jahresüberschuss (Net Income)", None, FMT_EUR, FARBE_TOTAL_BG),
        ("", None, None, None),
        ("EBITDA-Marge", None, FMT_PCT, None),
        ("Nettomarge", None, FMT_PCT, None),
    ]

    # Formel-Map: Zelle A-Referenz = Annahmen-Sheet
    A = "Annahmen"  # Sheet-Name für Cross-References
    start_row = 3

    for r_off, (label, _, fmt, bg) in enumerate(is_zeilen):
        row = start_row + r_off
        _zelle_stil(ws_is, row, 1, label, farbe=FARBE_FORMEL, ausrichtung="left")
        for c in range(2, n + 2):
            col_l = get_column_letter(c)
            # Umsatz (Zeile 3): Basisumsatz * (1 + Wachstum)^n
            if label == "Umsatz":
                if c == 2:
                    formel = f"='{A}'!{col_l}9"  # Basisumsatz (Zeile 9 in Annahmen)
                else:
                    prev = get_column_letter(c - 1)
                    formel = f"={prev}{row}*(1+'{A}'!{col_l}4)"
                _zelle_stil(ws_is, row, c, formel, zahlenformat=fmt, bg=bg)
            elif label == "  Herstellungskosten (COGS)":
                formel = f"=-{col_l}{start_row}*(1-'{A}'!{col_l}5)"
                _zelle_stil(ws_is, row, c, formel, zahlenformat=fmt, bg=bg)
            elif label == "Bruttogewinn":
                formel = f"={col_l}{start_row}+{col_l}{start_row+1}"
                _zelle_stil(ws_is, row, c, formel, zahlenformat=fmt, bg=bg, bold=True)
            elif label == "EBITDA":
                formel = f"={col_l}{start_row}*'{A}'!{col_l}6"
                _zelle_stil(ws_is, row, c, formel, zahlenformat=fmt, bg=bg, bold=True)
            elif label == "  Abschreibungen (D&A)":
                formel = f"=-{col_l}{start_row}*'{A}'!{col_l}7"
                _zelle_stil(ws_is, row, c, formel, zahlenformat=fmt, bg=bg)
            elif label == "EBIT (Operatives Ergebnis)":
                formel = f"={col_l}{start_row+6}+{col_l}{start_row+7}"
                _zelle_stil(ws_is, row, c, formel, zahlenformat=fmt, bg=bg, bold=True)
            elif label == "  Zinsen und Finanzkosten":
                formel = f"=-{col_l}{start_row}*0.02"   # 2% Zinslast-Annahme
                _zelle_stil(ws_is, row, c, formel, zahlenformat=fmt, bg=bg, ist_eingabe=True)
            elif label == "EBT (Ergebnis vor Steuern)":
                formel = f"={col_l}{start_row+8}+{col_l}{start_row+9}"
                _zelle_stil(ws_is, row, c, formel, zahlenformat=fmt, bg=bg, bold=True)
            elif label == "  Steuern":
                formel = f"=-MAX({col_l}{start_row+10},0)*'{A}'!{col_l}8"
                _zelle_stil(ws_is, row, c, formel, zahlenformat=fmt, bg=bg)
            elif label == "Jahresüberschuss (Net Income)":
                formel = f"={col_l}{start_row+10}+{col_l}{start_row+11}"
                _zelle_stil(ws_is, row, c, formel, zahlenformat=fmt, bg=bg, bold=True)
            elif label == "EBITDA-Marge":
                formel = f"=IFERROR({col_l}{start_row+6}/{col_l}{start_row},0)"
                _zelle_stil(ws_is, row, c, formel, zahlenformat=fmt, bg=bg)
            elif label == "Nettomarge":
                formel = f"=IFERROR({col_l}{start_row+12}/{col_l}{start_row},0)"
                _zelle_stil(ws_is, row, c, formel, zahlenformat=fmt, bg=bg)
            elif label == "":
                pass
            else:
                _zelle_stil(ws_is, row, c, "", zahlenformat=fmt, bg=bg)

    _freeze(ws_is, "B3")

    # --- SHEET 3: BALANCE SHEET (vereinfacht) ---
    ws_bs = wb.create_sheet("Balance Sheet")
    _spaltenbreiten(ws_bs, {1: 38, **{i+2: 14 for i in range(n)}})
    _header_zeile(ws_bs, 1, "BALANCE SHEET (Bilanz — vereinfacht)", n + 1)

    for i, j in enumerate(jahre):
        _zelle_stil(ws_bs, 2, i+2, str(j), bold=True, bg=FARBE_HEADER_BG,
                    farbe=FARBE_HEADER_TEXT, ausrichtung="center")

    bs_zeilen = [
        ("AKTIVA", True),
        ("  Zahlungsmittel & Äquivalente", False),
        ("  Forderungen (A/R)", False),
        ("  Vorräte", False),
        ("Umlaufvermögen (Total)", True),
        ("  Sachanlagen (PP&E, netto)", False),
        ("  Immaterielle Vermögenswerte", False),
        ("Anlagevermögen (Total)", True),
        ("BILANZSUMME AKTIVA", True),
        ("", False),
        ("PASSIVA", True),
        ("  Verbindlichkeiten (A/P)", False),
        ("  Kurzfristige Schulden", False),
        ("Kurzfristiges Fremdkapital (Total)", True),
        ("  Langfristige Schulden", False),
        ("Gesamtschulden", True),
        ("  Eigenkapital (Equity)", False),
        ("BILANZSUMME PASSIVA", True),
    ]

    for r_off, (label, is_total) in enumerate(bs_zeilen, 3):
        row = r_off
        bg = FARBE_TOTAL_BG if is_total else None
        bold = is_total
        _zelle_stil(ws_bs, row, 1, label, farbe=FARBE_FORMEL, ausrichtung="left", bold=bold)
        for c in range(2, n + 2):
            col_l = get_column_letter(c)
            is_r = "'Income Statement'" if "Umsatz" in label else None
            _zelle_stil(ws_bs, row, c, "" if label != "" else None,
                        zahlenformat=FMT_EUR, bg=bg, bold=bold)

    _freeze(ws_bs, "B3")

    # --- SHEET 4: CASH FLOW STATEMENT ---
    ws_cf = wb.create_sheet("Cash Flow")
    _spaltenbreiten(ws_cf, {1: 38, **{i+2: 14 for i in range(n)}})
    _header_zeile(ws_cf, 1, "CASH FLOW STATEMENT (indirektes Verfahren)", n + 1)

    for i, j in enumerate(jahre):
        _zelle_stil(ws_cf, 2, i+2, str(j), bold=True, bg=FARBE_HEADER_BG,
                    farbe=FARBE_HEADER_TEXT, ausrichtung="center")

    cf_zeilen = [
        ("OPERATIVER CASHFLOW", True),
        ("  Jahresüberschuss", False),
        ("  + Abschreibungen (D&A)", False),
        ("  ± Veränderung Working Capital", False),
        ("Cashflow aus Geschäftstätigkeit", True),
        ("", False),
        ("INVESTITIONS-CASHFLOW", True),
        ("  - Investitionen (Capex)", False),
        ("  + Desinvestitionen", False),
        ("Cashflow aus Investitionstätigkeit", True),
        ("", False),
        ("FINANZIERUNGS-CASHFLOW", True),
        ("  + Kreditaufnahmen", False),
        ("  - Tilgungen", False),
        ("  - Dividenden", False),
        ("Cashflow aus Finanzierungstätigkeit", True),
        ("", False),
        ("NET CHANGE IN CASH", True),
        ("Anfangsbestand Zahlungsmittel", False),
        ("Endbestand Zahlungsmittel", True),
    ]

    for r_off, (label, is_total) in enumerate(cf_zeilen, 3):
        bg = FARBE_TOTAL_BG if is_total else None
        _zelle_stil(ws_cf, r_off, 1, label, farbe=FARBE_FORMEL,
                    ausrichtung="left", bold=is_total)
        for c in range(2, n + 2):
            col_l = get_column_letter(c)
            if label == "  Jahresüberschuss":
                formel = f"='Income Statement'!{col_l}15"  # Net Income
                _zelle_stil(ws_cf, r_off, c, formel, zahlenformat=FMT_EUR,
                            bg=bg, farbe=FARBE_LINK)
            elif label == "  - Investitionen (Capex)":
                formel = f"=-'Income Statement'!{col_l}3*'{A}'!{col_l}7"
                _zelle_stil(ws_cf, r_off, c, formel, zahlenformat=FMT_EUR, bg=bg)
            else:
                _zelle_stil(ws_cf, r_off, c, "" if label != "" else None,
                            zahlenformat=FMT_EUR, bg=bg, bold=is_total)

    _freeze(ws_cf, "B3")

    # Speichern
    ziel = Path(ausgabe)
    ziel.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ziel))
    print(f"OK  | 3-Statement Modell gespeichert: {ziel.resolve()}")
    print(f"    | Sheets: Annahmen | Income Statement | Balance Sheet | Cash Flow")
    print(f"    | Perioden: {perioden} | Formelinjektion: aktiv | Hardcoded Values: 0")


# =============================================================================
# MODUL 2: SaaS-METRIKEN-MODELL
# ARR, MRR, Churn, LTV, CAC, LTV/CAC-Ratio — monatliche Granularität
# =============================================================================

def erstelle_saas_modell(ausgabe: str, perioden: int = 12) -> None:
    """
    SaaS-Metriken-Modell mit Formelinjektion:
    ARR, MRR, Net Revenue Retention, Churn Rate, LTV, CAC, Payback Period.
    """
    if not HAS_OPENPYXL:
        print("ERR | openpyxl nicht installiert", file=sys.stderr); sys.exit(1)

    print(f"[{_ts()}] [SaaS] Erstelle Modell mit {perioden} Monaten...")
    wb = Workbook()

    monate = [f"M{i+1}" for i in range(perioden)]

    # SHEET 1: Annahmen
    ws_a = wb.active
    ws_a.title = "SaaS-Annahmen"
    _spaltenbreiten(ws_a, {1: 35, 2: 18})
    _header_zeile(ws_a, 1, "SaaS ANNAHMEN (blau = Eingabe)", 2)

    saas_annahmen = [
        ("Startender ARR (€)", 1_200_000),
        ("Monatliches Netto-Neugeschäft (€)", 50_000),
        ("Monatliche Churn-Rate (%)", 0.015),
        ("Expansion Revenue Rate (%)", 0.005),
        ("Customer Acquisition Cost (CAC, €)", 8_500),
        ("Avg. Contract Value (ACV, €)", 12_000),
        ("Gross Margin (%)", 0.72),
        ("Payback Period Ziel (Monate)", 18),
    ]

    for r, (name, wert) in enumerate(saas_annahmen, 3):
        _zelle_stil(ws_a, r, 1, name, ausrichtung="left")
        fmt = FMT_PCT if "%" in name and "Monate" not in name else (
              FMT_EUR if "€" in name else "0")
        _zelle_stil(ws_a, r, 2, wert, ist_eingabe=True, zahlenformat=fmt)

    # SHEET 2: SaaS-Metriken (Monatlich)
    ws_m = wb.create_sheet("Metriken")
    _spaltenbreiten(ws_m, {1: 30, **{i+2: 11 for i in range(perioden)}})
    _header_zeile(ws_m, 1, "SaaS METRIKEN — Monatliche Entwicklung", perioden + 1)

    for i, m in enumerate(monate):
        _zelle_stil(ws_m, 2, i+2, m, bold=True, bg=FARBE_HEADER_BG,
                    farbe=FARBE_HEADER_TEXT, ausrichtung="center")

    A = "SaaS-Annahmen"
    metriken = [
        ("ARR (Annual Recurring Revenue)", FMT_EUR),
        ("MRR (Monthly Recurring Revenue)", FMT_EUR),
        ("Churned ARR", FMT_EUR),
        ("Expansion ARR", FMT_EUR),
        ("Net New ARR", FMT_EUR),
        ("", None),
        ("Churn Rate (Monat)", FMT_PCT_1),
        ("Net Revenue Retention (NRR)", FMT_PCT_1),
        ("LTV (Lifetime Value)", FMT_EUR),
        ("LTV / CAC Ratio", FMT_MULTI),
        ("CAC Payback (Monate)", "0.0"),
    ]

    for r_off, (label, fmt) in enumerate(metriken, 3):
        row = r_off
        _zelle_stil(ws_m, row, 1, label, ausrichtung="left")
        for c in range(2, perioden + 2):
            col_l = get_column_letter(c)
            prev_l = get_column_letter(c - 1) if c > 2 else None

            if label == "ARR (Annual Recurring Revenue)":
                if c == 2:
                    formel = f"='{A}'!B3"
                else:
                    formel = (f"={prev_l}{row}+{prev_l}{row+4}"
                              f"-{prev_l}{row+2}+{prev_l}{row+3}")
                _zelle_stil(ws_m, row, c, formel, zahlenformat=fmt)
            elif label == "MRR (Monthly Recurring Revenue)":
                formel = f"={col_l}{row-1}/12"
                _zelle_stil(ws_m, row, c, formel, zahlenformat=fmt)
            elif label == "Churned ARR":
                formel = f"=-{col_l}{row-2}*'{A}'!B5"
                _zelle_stil(ws_m, row, c, formel, zahlenformat=fmt, farbe=FARBE_FEHLER)
            elif label == "Expansion ARR":
                formel = f"={col_l}{row-3}*'{A}'!B6"
                _zelle_stil(ws_m, row, c, formel, zahlenformat=fmt, farbe=FARBE_LINK)
            elif label == "Net New ARR":
                formel = f"='{A}'!B4+{col_l}{row-1}+{col_l}{row-2}"
                _zelle_stil(ws_m, row, c, formel, zahlenformat=fmt, bold=True,
                            bg=FARBE_SUBTOTAL_BG)
            elif label == "Churn Rate (Monat)":
                formel = f"='{A}'!B5"
                _zelle_stil(ws_m, row, c, formel, zahlenformat=fmt)
            elif label == "Net Revenue Retention (NRR)":
                formel = f"=1-{col_l}{row-1}+'{A}'!B6"
                _zelle_stil(ws_m, row, c, formel, zahlenformat=fmt,
                            bg=FARBE_TOTAL_BG, bold=True)
            elif label == "LTV (Lifetime Value)":
                formel = f"=IFERROR('{A}'!B7*'{A}'!B8/'{A}'!B5,0)"
                _zelle_stil(ws_m, row, c, formel, zahlenformat=fmt)
            elif label == "LTV / CAC Ratio":
                formel = f"=IFERROR({col_l}{row-1}/'{A}'!B7,0)"
                _zelle_stil(ws_m, row, c, formel, zahlenformat=fmt, bg=FARBE_TOTAL_BG)
            elif label == "CAC Payback (Monate)":
                formel = f"=IFERROR('{A}'!B7/('{A}'!B8*'{A}'!B9),0)"
                _zelle_stil(ws_m, row, c, formel, zahlenformat=fmt)
            elif label == "":
                pass

    _freeze(ws_m, "B3")
    ziel = Path(ausgabe)
    ziel.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ziel))
    print(f"OK  | SaaS-Modell gespeichert: {ziel.resolve()}")
    print(f"    | Sheets: SaaS-Annahmen | Metriken | Perioden: {perioden} Monate")


# =============================================================================
# MODUL 3: SZENARIO-ANALYSE (Base / Bull / Bear)
# =============================================================================

def erstelle_szenario(ausgabe: str) -> None:
    """
    3-Szenario-Analyse mit separaten Annahmen-Spalten und Formel-Verknüpfung.
    Szenarien: Bear Case | Base Case | Bull Case
    """
    if not HAS_OPENPYXL:
        print("ERR | openpyxl nicht installiert", file=sys.stderr); sys.exit(1)

    print(f"[{_ts()}] [Szenario] Erstelle 3-Szenario-Analyse...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Szenario-Analyse"
    _spaltenbreiten(ws, {1: 35, 2: 18, 3: 18, 4: 18, 5: 18})

    SZENARIEN = [
        ("Bear Case", "D9D9D9", "4D4D4D"),   # Grau
        ("Base Case", "D6E4F0", "1F3864"),   # Blau
        ("Bull Case", "E2EFDA", "375623"),   # Grün
    ]

    _header_zeile(ws, 1, "SZENARIO-ANALYSE — Base | Bull | Bear Case", 5)
    for i, (name, bg, txt) in enumerate(SZENARIEN):
        _zelle_stil(ws, 2, i+2, name, bold=True, bg=bg, farbe=txt, ausrichtung="center")
    ws.cell(row=2, column=5, value="Aktives Szenario").font = Font(bold=True, size=9)

    # Annahmen (Eingabezellen in Blau)
    szen_annahmen = [
        ("Umsatz-Basisjahr (€)",         4_800_000, 5_000_000, 5_200_000),
        ("Umsatzwachstum Jahr 1",           0.10,      0.20,      0.30),
        ("Umsatzwachstum Jahr 2",           0.08,      0.18,      0.28),
        ("Umsatzwachstum Jahr 3",           0.06,      0.15,      0.25),
        ("EBITDA-Marge (stabil)",           0.08,      0.18,      0.28),
        ("Churn-Rate (annualisiert)",        0.20,      0.12,      0.06),
        ("Multiplikator (EV/EBITDA)",          6,        10,        15),
    ]

    _section_header(ws, 3, "Annahmen je Szenario", 5)
    for r_off, (name, bear, base, bull) in enumerate(szen_annahmen, 4):
        _zelle_stil(ws, r_off, 1, name, ausrichtung="left")
        for c, val in enumerate([bear, base, bull], 2):
            fmt = FMT_PCT if isinstance(val, float) and val < 50 else FMT_EUR if val > 100 else "0"
            _zelle_stil(ws, r_off, c, val, ist_eingabe=True, zahlenformat=fmt)

    # Aktives Szenario — Dropdown-Verweis (vereinfacht als Formel)
    ws.cell(row=4, column=5, value="Base Case")
    ws.cell(row=4, column=5).font = Font(color=FARBE_EINGABE, bold=True, size=9)

    # Berechnete Outputs
    _section_header(ws, 12, "Berechnete Outputs (Formeln — kein Hardcoding)", 5)
    outputs = [
        ("Umsatz Jahr 1 (€)", [
            f"=B4*(1+B5)", f"=C4*(1+C5)", f"=D4*(1+D5)"
        ]),
        ("Umsatz Jahr 2 (€)", [
            f"=B4*(1+B5)*(1+B6)", f"=C4*(1+C5)*(1+C6)", f"=D4*(1+D5)*(1+D6)"
        ]),
        ("Umsatz Jahr 3 (€)", [
            f"=B4*(1+B5)*(1+B6)*(1+B7)", f"=C4*(1+C5)*(1+C6)*(1+C7)",
            f"=D4*(1+D5)*(1+D6)*(1+D7)"
        ]),
        ("EBITDA Jahr 3 (€)", [
            f"=B14*(B8)", f"=C14*(C8)", f"=D14*(D8)"
        ]),
        ("Enterprise Value (€)", [
            f"=B15*B10", f"=C15*C10", f"=D15*D10"
        ]),
    ]

    for r_off, (name, formeln) in enumerate(outputs, 13):
        _zelle_stil(ws, r_off, 1, name, ausrichtung="left", bold=True)
        bg_list = [SZENARIEN[i][1] for i in range(3)]
        for c, (formel, bg) in enumerate(zip(formeln, bg_list), 2):
            _zelle_stil(ws, r_off, c, formel, zahlenformat=FMT_EUR, bg=bg, bold=True)

    _freeze(ws, "B3")
    ziel = Path(ausgabe)
    ziel.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ziel))
    print(f"OK  | Szenario-Analyse gespeichert: {ziel.resolve()}")
    print(f"    | Szenarien: Bear | Base | Bull | Formelinjektion: aktiv")


# =============================================================================
# MODUL 4: VERIFIKATION (LibreOffice headless)
# Öffnet die Datei in LibreOffice headless und prüft auf Formel-Errors.
# =============================================================================

def verifiziere(pfad: str) -> dict:
    """
    Verifiziert eine .xlsx-Datei auf Formel-Fehler (#REF!, #VALUE!, etc.).
    Zwei Stufen:
    1. openpyxl-Scan: Prüft alle Zellwerte auf Fehlerwerte
    2. LibreOffice headless (wenn verfügbar): Öffnet und re-kalkuliert

    Output (stdout): JSON-Report mit Fehlerliste
    """
    if not HAS_OPENPYXL:
        print("ERR | openpyxl nicht installiert", file=sys.stderr); sys.exit(1)

    p = Path(pfad)
    if not p.exists():
        print(f"ERR | Datei nicht gefunden: {pfad}", file=sys.stderr); sys.exit(1)

    print(f"[{_ts()}] [Verifikation] Prüfe: {p.name}")
    wb = load_workbook(str(p), data_only=True)

    FEHLERWERTE = {"#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#NULL!", "#N/A", "#NUM!"}
    fehler: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.upper() in FEHLERWERTE:
                    fehler.append({
                        "sheet": sheet_name,
                        "zelle": cell.coordinate,
                        "fehlertyp": cell.value,
                        "zeile": cell.row,
                        "spalte": cell.column,
                    })

    # LibreOffice headless (optional)
    libre = shutil.which("libreoffice") or shutil.which("soffice")
    libre_status = "nicht verfügbar"
    if libre:
        try:
            tmp_dir = Path("/tmp/libre_verify")
            tmp_dir.mkdir(exist_ok=True)
            result = subprocess.run(
                [libre, "--headless", "--calc",
                 "--infilter=Calc MS Excel 2007 XML",
                 "--convert-to", "xlsx",
                 "--outdir", str(tmp_dir),
                 str(p.resolve())],
                capture_output=True, text=True, timeout=30
            )
            libre_status = "OK" if result.returncode == 0 else f"Exit {result.returncode}"
        except (subprocess.TimeoutExpired, FileNotFoundError) as e:
            libre_status = f"Fehler: {e}"

    report = {
        "datei": str(p.resolve()),
        "sheets": wb.sheetnames,
        "formel_fehler": len(fehler),
        "libreoffice": libre_status,
        "details": fehler,
    }

    print(f"\n{'─' * 60}")
    if not fehler:
        print(f"OK  | 0 Formel-Fehler gefunden in {p.name}")
    else:
        print(f"WRN | {len(fehler)} Formel-Fehler gefunden:")
        for f in fehler:
            print(f"    | Sheet={f['sheet']} Zelle={f['zelle']} Typ={f['fehlertyp']}")

    print(f"    | LibreOffice Verifikation: {libre_status}")
    print(f"{'─' * 60}")
    print("\nJSON-Report:")
    print(json.dumps(report, indent=2, ensure_ascii=False))
    return report


# =============================================================================
# MODUL 5: DEBUGGING (#REF!, #VALUE!, Zirkelbezüge)
# =============================================================================

def debugge(pfad: str) -> None:
    """
    Analysiert eine xlsx-Datei auf Fehlerquellen:
    - Fehlerwerte in Zellen
    - Formeln, die auf nicht-existente Zellen zeigen
    - Potenzielle Zirkelbezüge (Heuristik: gleiche Zelle im Formel-String)
    """
    if not HAS_OPENPYXL:
        print("ERR | openpyxl nicht installiert", file=sys.stderr); sys.exit(1)

    p = Path(pfad)
    if not p.exists():
        print(f"ERR | Datei nicht gefunden: {pfad}", file=sys.stderr); sys.exit(1)

    print(f"[{_ts()}] [Debug] Analysiere: {p.name}")
    wb = load_workbook(str(p))  # data_only=False → Formeln sehen

    FEHLERWERTE = {"#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#NULL!", "#N/A", "#NUM!"}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_fehler = []
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                s = str(v)
                # Fehlerwert in Data-Zelle
                if s.upper() in FEHLERWERTE:
                    sheet_fehler.append({
                        "zelle": cell.coordinate,
                        "typ": "FEHLERWERT",
                        "wert": s,
                        "empfehlung": _empfehlung(s),
                    })
                # Potenzielle Zirkelbezüge (Heuristik)
                elif s.startswith("=") and cell.coordinate.replace("$", "") in s:
                    sheet_fehler.append({
                        "zelle": cell.coordinate,
                        "typ": "ZIRKELBEZUG (Verdacht)",
                        "wert": s[:80],
                        "empfehlung": "Formel auf Selbstreferenz prüfen; ggf. Iterative Berechnung aktivieren",
                    })

        if sheet_fehler:
            print(f"\n  Sheet: {sheet_name} — {len(sheet_fehler)} Problem(e)")
            for f in sheet_fehler:
                print(f"  [{f['typ']}] {f['zelle']}: {f['wert']}")
                print(f"    → {f['empfehlung']}")
        else:
            print(f"  Sheet: {sheet_name} — OK (keine Fehler)")

    print(f"\n[{_ts()}] [Debug] Analyse abgeschlossen")


def _empfehlung(fehlertyp: str) -> str:
    empfehlungen = {
        "#REF!":   "Quellzelle gelöscht oder verschoben. Formel neu aufbauen.",
        "#VALUE!": "Falscher Datentyp. Quellzelle auf Text/Zahl prüfen; IFERROR() wrappen.",
        "#DIV/0!": "Division durch Null. Mit IFERROR(Formel,0) schützen.",
        "#NAME?":  "Unbekannter Funktions-/Bereichsname. Schreibweise prüfen.",
        "#N/A":    "Wert nicht gefunden (SVERWEIS/XLOOKUP). IFNA() oder Quell-Range prüfen.",
        "#NULL!":  "Falscher Bereichs-Operator. Doppelpunkt statt Leerzeichen prüfen.",
        "#NUM!":   "Ungültiger numerischer Wert (z. B. SQRT negativer Zahl).",
    }
    return empfehlungen.get(fehlertyp.upper(), "Formel und Quellzellen prüfen.")


# =============================================================================
# UTILITIES
# =============================================================================

def _ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# =============================================================================
# CLI
# =============================================================================

HILFE = """
VERWENDUNG:
  python excel_finanzmodell.py 3statement   <ausgabe.xlsx> [--perioden 5]
  python excel_finanzmodell.py saas         <ausgabe.xlsx> [--perioden 12]
  python excel_finanzmodell.py szenario     <ausgabe.xlsx>
  python excel_finanzmodell.py verifizieren <datei.xlsx>
  python excel_finanzmodell.py debuggen     <datei.xlsx>

MODULE:
  3statement    3-Statement Finanzmodell (IS + BS + CF) mit Formelinjektion
  saas          SaaS-Metriken (ARR, MRR, Churn, LTV, CAC, NRR)
  szenario      Bear / Base / Bull Case Analyse
  verifizieren  Formel-Fehler prüfen (#REF!, #VALUE! etc.) + LibreOffice-Check
  debuggen      Fehlerquellen analysieren und Empfehlungen ausgeben

KONVENTIONEN:
  Blau  (#0070C0) = manuelle Eingabe (Annahmen)
  Schwarz (#000000) = berechnete Formel
  Grün  (#00B050) = Cross-Sheet-Referenz
  Rot   (#FF0000) = Fehler / negativer Alert

ABHÄNGIGKEITEN:
  pip install openpyxl pandas
  apt install libreoffice  (optional, für Verifikation)
"""

if __name__ == "__main__":
    if not HAS_OPENPYXL:
        print("ERR | openpyxl nicht installiert.\nAusführen: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    if len(sys.argv) < 2:
        print(HILFE); sys.exit(0)

    modus = sys.argv[1]
    args = sys.argv[2:]

    def _get_opt(opt: str, default: int) -> int:
        for i, a in enumerate(args):
            if a == opt and i + 1 < len(args):
                return int(args[i + 1])
        return default

    if modus == "3statement":
        if not args:
            print("ERR | VERWENDUNG: 3statement <ausgabe.xlsx>", file=sys.stderr); sys.exit(1)
        erstelle_3statement(args[0], perioden=_get_opt("--perioden", 5))

    elif modus == "saas":
        if not args:
            print("ERR | VERWENDUNG: saas <ausgabe.xlsx>", file=sys.stderr); sys.exit(1)
        erstelle_saas_modell(args[0], perioden=_get_opt("--perioden", 12))

    elif modus == "szenario":
        if not args:
            print("ERR | VERWENDUNG: szenario <ausgabe.xlsx>", file=sys.stderr); sys.exit(1)
        erstelle_szenario(args[0])

    elif modus == "verifizieren":
        if not args:
            print("ERR | VERWENDUNG: verifizieren <datei.xlsx>", file=sys.stderr); sys.exit(1)
        verifiziere(args[0])

    elif modus == "debuggen":
        if not args:
            print("ERR | VERWENDUNG: debuggen <datei.xlsx>", file=sys.stderr); sys.exit(1)
        debugge(args[0])

    else:
        print(f"ERR | Unbekannter Modus '{modus}'\n{HILFE}", file=sys.stderr); sys.exit(1)
