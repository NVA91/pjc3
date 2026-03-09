#!/usr/bin/env python3
# LADESTUFE 3 — RESSOURCE
# Quellcode belastet das Kontextfenster nicht. Nur stdout (Ergebnis) wird
# dem Modell zurückgegeben. Ausführung in Sandbox auf expliziten Nutzerruf.
#
# BESCHREIBUNG: Erstellt eine Excel-Aufgabentabelle (.xlsx) mit farbigen
#               Status-Badges, gefrorenem Header und Auto-Filter.
# VERWENDUNG:   python excel_aufgaben.py <ausgabe.xlsx>
# ABHÄNGIGKEIT: pip install openpyxl

import sys
from pathlib import Path
from datetime import date

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("FEHLER: openpyxl nicht installiert. Ausführen: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# --- Konfiguration ---
HEADERS = ["ID", "Aufgabe", "Priorität", "Status", "Fällig am", "Verantwortlich", "Notizen"]
SPALTEN_BREITEN = [5, 40, 12, 15, 12, 20, 30]
ZEILENHÖHE = 22

# Status-Farbschema (unveränderlich — laut SKILL.md festgelegt)
STATUS_FARBEN = {
    "Offen":     "FFE699",  # Gelb
    "In Arbeit": "9DC3E6",  # Blau
    "Erledigt":  "A9D18E",  # Grün
    "Blockiert": "FF7C80",  # Rot
}

HEADER_FARBE = "2F5496"   # Dunkelblau
HEADER_SCHRIFT = "FFFFFF"  # Weiß


def _rand_duenn() -> Border:
    seite = Side(style="thin", color="D1D9E6")
    return Border(left=seite, right=seite, top=seite, bottom=seite)


def _erstelle_header(ws, rand: Border) -> None:
    header_fill = PatternFill(start_color=HEADER_FARBE, end_color=HEADER_FARBE, fill_type="solid")
    header_font = Font(color=HEADER_SCHRIFT, bold=True, size=10)

    for col, (titel, breite) in enumerate(zip(HEADERS, SPALTEN_BREITEN), 1):
        zelle = ws.cell(row=1, column=col, value=titel)
        zelle.fill = header_fill
        zelle.font = header_font
        zelle.alignment = Alignment(horizontal="center", vertical="center")
        zelle.border = rand
        ws.column_dimensions[get_column_letter(col)].width = breite

    ws.row_dimensions[1].height = ZEILENHÖHE + 4


def _schreibe_zeile(ws, row_idx: int, aufgabe: list, rand: Border) -> None:
    for col_idx, wert in enumerate(aufgabe, 1):
        zelle = ws.cell(row=row_idx, column=col_idx, value=wert)
        zelle.border = rand
        zelle.alignment = Alignment(vertical="center", wrap_text=True)
        zelle.font = Font(size=10)

    ws.row_dimensions[row_idx].height = ZEILENHÖHE

    # Status-Badge einfärben (Spalte 4)
    status = aufgabe[3] if len(aufgabe) > 3 else None
    if status and status in STATUS_FARBEN:
        farbe = STATUS_FARBEN[status]
        ws.cell(row=row_idx, column=4).fill = PatternFill(
            start_color=farbe, end_color=farbe, fill_type="solid"
        )


def erstelle_aufgaben_tabelle(ausgabe_pfad: str, aufgaben: list[list] | None = None) -> None:
    """
    Erstellt eine Excel-Aufgabentabelle.

    Args:
        ausgabe_pfad: Zielpfad für die .xlsx-Datei.
        aufgaben: Liste von Zeilen [ID, Aufgabe, Priorität, Status, Fällig, Verantwortlich, Notizen].
                  Wird None übergeben, werden Beispieldaten eingefügt.
    """
    if aufgaben is None:
        aufgaben = [
            [1, "Projektplan erstellen",    "Hoch",   "Offen",     date.today(), "Max M.",  ""],
            [2, "README schreiben",         "Mittel", "In Arbeit", date.today(), "Anna L.", "Entwurf vorhanden"],
            [3, "Tests implementieren",     "Hoch",   "Offen",     date.today(), "Max M.",  ""],
            [4, "Deployment vorbereiten",   "Niedrig","Blockiert", date.today(), "Anna L.", "Wartet auf Freigabe"],
            [5, "Code-Review durchführen",  "Mittel", "Erledigt",  date.today(), "Team",    ""],
        ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Aufgaben"
    rand = _rand_duenn()

    _erstelle_header(ws, rand)

    for row_idx, aufgabe in enumerate(aufgaben, 2):
        _schreibe_zeile(ws, row_idx, aufgabe, rand)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

    # Ausgabe sichern
    ziel = Path(ausgabe_pfad)
    ziel.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ziel))

    print(f"OK | Tabelle gespeichert: {ziel.resolve()}")
    print(f"   | Spalten: {len(HEADERS)} | Zeilen (ohne Header): {len(aufgaben)}")
    print(f"   | Status-Typen: {', '.join(STATUS_FARBEN.keys())}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("VERWENDUNG: python excel_aufgaben.py <ausgabe.xlsx>", file=sys.stderr)
        sys.exit(1)

    ausgabe = sys.argv[1]
    if not ausgabe.endswith(".xlsx"):
        print("FEHLER: Ausgabedatei muss auf .xlsx enden.", file=sys.stderr)
        sys.exit(1)

    erstelle_aufgaben_tabelle(ausgabe)
